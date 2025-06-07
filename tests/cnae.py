"""test file para ler o cnae do excel do ibge"""

from pathlib import Path

from openpyxl import load_workbook


def read_cnae(file_path: Path) -> dict:
    """Lê o arquivo Excel do IBGE e retorna um dicionário com os dados do CNAE.

    Args:
        file_path (Path): Caminho para o arquivo Excel.

    Returns:
        dict: Dicionário contendo os dados do CNAE.
    """
    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb["Estrutura Det. CNAE Subclass2.3"]
    cnae_dict = {}

    "Seção	Divisão	Grupo	Classe	Subclasse"
    section: str | None = None
    division: str | None = None
    group: str | None = None
    clazz: str | None = None
    subclass: str | None = None

    with open("cnae_orm.txt", "w", encoding="utf-8") as f:
        for row in ws.iter_rows(values_only=True):
            if row[0] and row[0] != section:
                section = str(row[0])
                print(
                    (
                        f"parentsection = EconomicActivitySection(code='{section}', "
                        f"name='{row[5]}')"
                    ),
                    file=f,
                )
            if row[1] and row[1] != division:
                division = str(row[1])
                print(
                    (
                        f"parentdivision = EconomicActivityDivision(code='{division}', "
                        f"name='{row[5]}', parent=parentsection)"
                    ),
                    file=f,
                )
            if row[2] and row[2] != group:
                group = str(row[2])
                print(
                    (
                        f"parentgroup = EconomicActivityGroup(code='{group}', "
                        f"name='{row[5]}', parent=parentdivision)"
                    ),
                    file=f,
                )
            if row[3] and row[3] != clazz:
                clazz = str(row[3])
                print(
                    (
                        f"parentclazz = EconomicActivityClass(code='{clazz}', "
                        f"name='{row[5]}', parent=parentgroup)"
                    ),
                    file=f,
                )
            if row[4] and row[4] != subclass:
                subclass = str(row[4])
                print(
                    (
                        f"parentsubclass = EconomicActivitySubclass(code='{subclass}', "
                        f"name='{row[5]}', parent=parentclazz)"
                    ),
                    file=f,
                )
            cnae_code, description = row[0], row[5]
            cnae_dict[cnae_code] = description

    return cnae_dict


if __name__ == "__main__":
    # Exemplo de uso
    file_path = Path("tests/cnae.xlsx")  # Substitua pelo caminho correto do arquivo
    cnae_data = read_cnae(file_path)
    for code, desc in cnae_data.items():
        print(f"{code}: {desc}")
